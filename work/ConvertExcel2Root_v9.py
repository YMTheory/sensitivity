########################################################################
##
## Main script to convert values in summary Excel table into ROOT tree
## Usage: 
## $python ConvertExcel2Root.py
##
########################################################################


import ROOT
import openpyxl
import sys
import time

ROOT.gSystem.Load('../lib/libnEXOSensitivity.so')

######################################################
############### EXCEL TABLE READER ###################
######################################################
class ExcelTableReader:
    # This class carries all table info
    # Modify '__init__' appropriately for table in use  

    def __init__(self,inTableName):
        self.filename = inTableName #'../tables/Summary_v68_2016-06-21_0nu.xlsx' #'../tables/Summary_v62_2016-06-04_0nu_tpc_elec.xlsx'
        self.wb = openpyxl.load_workbook(filename=self.filename,read_only=True,data_only=True)

        self.quantile = 1.64

        self.suffixes = ['SS','MS']

        self.specific_activity_sheet = {'name': 'SpecificActivities',\
                                        'startRow': 2,\
                                        'componentColumn': 'A',\
                                        'isotopeColumn': 'B',\
                                        'specColumn': 'H',\
                                        'specErrColumn': 'I',\
                                        'activColumn': 'J',\
                                        'activErrColumn': 'K',\
                                        'idColumn': 'L'}

        self.counts_sheet = {'name': '%s_ExpectedCounts',\
                             'startRow': 6,\
                             'componentColumn': 'A',\
                             'isotopeColumn': 'B',\
                             'cvColumn': {'SS': 'CQ', 'MS': 'CQ'},\
                             'errorColumn': {'SS': 'CR', 'MS': 'CR'},\
                             'ulColumn': {'SS': 'CV', 'MS': 'CV'},\
                             'llColumn': {'SS': 'CU', 'MS': 'CU'}}

        self.hiteff_sheet = {'name': 'MC_RawCounts_%s_Integrals',\
                             'startRow': 6,\
                             'componentColumn': 'A',\
                             'isotopeColumn': 'B',\
                             'nColumn': 'D',\
                             'kColumn': 'AI'}

        self.hiteff_sheet_fwhm = {'1tColumn': 'G', 
                                  '3tColumn': 'O',\
                                  'fvColumn': 'S',\
                                  '0p5tColumn': 'E',\
                                  '1p5tColumn': 'I',\
                                  '2tColumn': 'K',\
                                  '2p5tColumn': 'M',\
                                  '3p5tColumn': 'Q'}

        self.hiteff_sheet_3sig = {'1tColumn': 'BC',\
                                  '3tColumn': 'BK',\
                                  'fvColumn': 'BO',\
                                  '0p5tColumn': 'BA',\
                                  '1p5tColumn': 'BE',\
                                  '2tColumn': 'BG',\
                                  '2p5tColumn': 'BI',\
                                  '3p5tColumn': 'BM'}

        self.hiteff_sheet_2sig = {'1tColumn': 'BS',\
                                  '3tColumn': 'CA',\
                                  'fvColumn': 'CE',\
                                  '0p5tColumn': 'BQ',\
                                  '1p5tColumn': 'BU',\
                                  '2tColumn': 'BW',\
                                  '2p5tColumn': 'BY',\
                                  '3p5tColumn': 'CC'}

        self.hiteff_sheet_1sig = {'1tColumn': 'CI',\
                                  '3tColumn': 'CQ',\
                                  'fvColumn': 'CU',\
                                  '0p5tColumn': 'CG',\
                                  '1p5tColumn': 'CK',\
                                  '2tColumn': 'CM',\
                                  '2p5tColumn': 'CO',\
                                  '3p5tColumn': 'CS'}

        self.constants_sheet = {'name': 'Constants',\
                                'startHLRow': 16,\
                                'endHLRow': 25,\
                                'isotopeColumn': 'A',\
                                'hlColumn': 'B'}

        self.name_dict = {}
        self.name_dict["OuterCryoResin"] = 'Outer Cryostat (Resin)'
        self.name_dict["OuterCryoFiber"] = 'Outer Cryostat (Fiber)'
        self.name_dict["OuterCryoSupportResin"] = 'Outer Cryostat Support (Resin)'
        self.name_dict["OuterCryoSupportFiber"] = 'Outer Cryostat Support (Fiber)'
        self.name_dict["InnerCryoResin"] = 'Inner Cryostat (Resin)'
        self.name_dict["InnerCryoFiber"] = 'Inner Cryostat (Fiber)'
        self.name_dict["InnerCryoSupportResin"] = 'Inner Cryostat Support (Resin)'
        self.name_dict["InnerCryoSupportFiber"] = 'Inner Cryostat Support (Fiber)'
        self.name_dict["InnerCryoLiner"] = 'Inner Cryostat Liner'
        self.name_dict["HFE"] = 'HFE'
        self.name_dict["HVTubes"] = 'HV Tubes'
        self.name_dict["HVCables"] = 'HV Cables'
        self.name_dict["HVFeedthru"] = 'HV Feedthrough'
        self.name_dict["HVFeedthruCore"] = 'HV Feedthrough Core'
        self.name_dict["HVPlunger"] = 'HV Plunger'
        self.name_dict["CalibrationGuideTube1"] = 'Calibration Guide Tube 1'
        self.name_dict["CalibrationGuideTube2"] = 'Calibration Guide Tube 2'
        self.name_dict["TPC"] = 'TPC Vessel'
        self.name_dict["TPCSupportCone"] = 'TPC Support Cone'
        self.name_dict["CathodeRn"] = 'Cathode (Radon)'
        self.name_dict["Cathode"] = 'Cathode'
        self.name_dict["Bulge"] = 'Bulge'
        self.name_dict["FieldRings"] = 'Field Rings'
        self.name_dict["SupportSpacers"] = 'Support Rods and Spacers'
        self.name_dict["SiPMStaves"] = 'SiPM Staves'
        self.name_dict["SiPMModule"] = 'SiPM Module (Interposer)'
        self.name_dict["SiPMElectronics"] = 'SiPM Electronics'
        self.name_dict["SiPMCables"] = 'SiPM Cables'
        self.name_dict["SiPMs"] = 'SiPMs'
        self.name_dict["ChargeTilesCables"] = 'Charge Tiles Cables'
        self.name_dict["ChargeTilesElectronics"] = 'Charge Tiles Electronics'
        self.name_dict["ChargeTilesSupport"] = 'Charge Tiles Support'
        self.name_dict["ChargeTilesBacking"] = 'Charge Tiles Backing'
        self.name_dict["LXe"] = 'Full LXe'
        #        self.name_dict["bb2n"] = 'Full LXe'
        #        self.name_dict["bb0n"] = 'Full LXe'
        self.name_dict["ActiveLXe"] = 'Active LXe'
        self.name_dict["InactiveLXe"] = 'Inactive LXe'
        self.name_dict["SolderAnode"] = 'Solder (Anode)'
        self.name_dict["SolderSiPM"] = 'Solder (SiPM)'
        self.ReadContents()

    def ReadContents(self):
        
        self.ReadComponents()
        self.ReadHalflife()
        self.ReadActivity()
        for suffix in self.suffixes:
            self.ReadCounts(suffix)
            self.ReadHitEfficiency(suffix)

        return        

    def ReadComponents(self):
        print('Reading components...')

        sheet = self.specific_activity_sheet
        ws = self.wb.get_sheet_by_name(sheet['name'])

        self.components = {}
        for cells in ws.iter_rows(row_offset=self.specific_activity_sheet['startRow']-1):
            if cells[0].value is None:
                break
            row = cells[0].row
            component = self.GetCellValue(ws,sheet['componentColumn'],row)
            print("##%s##" % component)
            isotope = self.GetCellValue(ws,sheet['isotopeColumn'],row)
            pdf = self.GetPdfName(component,isotope)
            
            self.components[pdf] = ROOT.ExcelTableValues(ROOT.TString(pdf),ROOT.TString(component),ROOT.TString(isotope))

        return

    def ReadHalflife(self):
        print('Reading half-life...')
        
        sheet = self.constants_sheet
        ws = self.wb.get_sheet_by_name(sheet['name'])

        start_row = sheet['startHLRow']
        end_row = sheet['endHLRow']
        
        for pdf in self.components:
            for row in range(start_row,end_row+1):
                isotope = self.GetCellValue(ws,sheet['isotopeColumn'],row)
                if self.components[pdf].fIsotope == isotope:
                    self.components[pdf].SetHalflife(self.GetCellValue(ws,sheet['hlColumn'],row))

        return

    def ReadActivity(self):
        print('Reading activity...')

        sheet = self.specific_activity_sheet
        ws = self.wb.get_sheet_by_name(sheet['name'])
        number_components = len(self.components)

        activ = {}
        start_row = sheet['startRow']
        end_row = sheet['startRow']+number_components
        for row in range(start_row,end_row):
            component = self.GetCellValue(ws,sheet['componentColumn'],row)
            isotope = self.GetCellValue(ws,sheet['isotopeColumn'],row)
            pdf = self.GetPdfName(component,isotope)
            print row, '/', end_row, ':', component, isotope, pdf

            activ[pdf] = {'spec':self.GetCellValue(ws,sheet['specColumn'],row),\
                          'specErr':self.GetCellValue(ws,sheet['specErrColumn'],row),\
                          'activ':self.GetCellValue(ws,sheet['activColumn'],row),\
                          'activErr':self.GetCellValue(ws,sheet['activErrColumn'],row),\
                          'id':self.GetCellValue(ws,sheet['idColumn'],row)}
            
            self.components[pdf].SetActivity( activ[pdf]['spec'],\
                                              activ[pdf]['specErr'],\
                                              activ[pdf]['activ'],\
                                              activ[pdf]['activErr'],\
                                              ROOT.TString(activ[pdf]['id']) ) 
        return         

    def ReadCounts(self,suffix):
        print 'Reading counts...', suffix

        sheet = self.counts_sheet
        ws = self.wb.get_sheet_by_name(sheet['name']%(suffix))
        number_components = len(self.components)

        counts = {}
        start_row = sheet['startRow']
        end_row = sheet['startRow']+number_components
        for row in range(start_row,end_row):
            component = self.GetCellValue(ws,sheet['componentColumn'],row)
            isotope = self.GetCellValue(ws,sheet['isotopeColumn'],row)
            pdf = self.GetPdfName(component,isotope)
            print row, '/', end_row, ':', component, isotope, pdf
            counts[pdf] = {'cv':self.GetCellValue(ws,sheet['cvColumn'][suffix],row),\
                           'error':self.GetCellValue(ws,sheet['errorColumn'][suffix],row),\
                           'ul':self.GetCellValue(ws,sheet['ulColumn'][suffix],row) }

            #counts[pdf]['ul'] = (self.quantile*counts[pdf]['error'],\
            #                      counts[pdf]['cv'] + self.quantile*counts[pdf]['error'])[counts[pdf]['cv'] > 0]
                       
            self.components[pdf].SetExpectedCounts(counts[pdf]['cv'],counts[pdf]['error'],counts[pdf]['ul'],suffix)
    
        return 

    def ReadHitEfficiency(self,suffix):
        print 'Reading hit efficiency...', suffix

        sheet = self.hiteff_sheet
        sheet_fwhm = self.hiteff_sheet_fwhm
        sheet_3sig = self.hiteff_sheet_3sig
        sheet_2sig = self.hiteff_sheet_2sig
        sheet_1sig = self.hiteff_sheet_1sig
        ws = self.wb.get_sheet_by_name(sheet['name']%(suffix))
        number_components = len(self.components)

        hiteff = {}
        start_row = sheet['startRow']
        end_row = sheet['startRow']+number_components
        for row in range(start_row,end_row):
            component = self.GetCellValue(ws,sheet['componentColumn'],row)
            isotope = self.GetCellValue(ws,sheet['isotopeColumn'],row)
            pdf = self.GetPdfName(component,isotope)
            print row, '/', end_row, ':', component, isotope, pdf
            hiteff[pdf] =  { 'n':self.GetCellValue(ws,sheet['nColumn'],row),\
                             'k':self.GetCellValue(ws,sheet['kColumn'],row) }

            if hiteff[pdf]['k'] == 0:
                hiteff[pdf]['fwhm_1t'] = 1
                hiteff[pdf]['fwhm_2t'] = 1
                hiteff[pdf]['fwhm_3t'] = 1
                hiteff[pdf]['fwhm_fv'] = 1
                hiteff[pdf]['fwhm_0p5t'] = 1
                hiteff[pdf]['fwhm_1p5t'] = 1
                hiteff[pdf]['fwhm_2p5t'] = 1
                hiteff[pdf]['fwhm_3p5t'] = 1
                hiteff[pdf]['3sig_1t'] = 1
                hiteff[pdf]['3sig_2t'] = 1
                hiteff[pdf]['3sig_3t'] = 1
                hiteff[pdf]['3sig_fv'] = 1
                hiteff[pdf]['3sig_0p5t'] = 1
                hiteff[pdf]['3sig_1p5t'] = 1
                hiteff[pdf]['3sig_2p5t'] = 1
                hiteff[pdf]['3sig_3p5t'] = 1
                hiteff[pdf]['2sig_1t'] = 1
                hiteff[pdf]['2sig_2t'] = 1
                hiteff[pdf]['2sig_3t'] = 1
                hiteff[pdf]['2sig_fv'] = 1
                hiteff[pdf]['2sig_0p5t'] = 1
                hiteff[pdf]['2sig_1p5t'] = 1
                hiteff[pdf]['2sig_2p5t'] = 1
                hiteff[pdf]['2sig_3p5t'] = 1
                hiteff[pdf]['1sig_1t'] = 1 
                hiteff[pdf]['1sig_2t'] = 1 
                hiteff[pdf]['1sig_3t'] = 1 
                hiteff[pdf]['1sig_fv'] = 1 
                hiteff[pdf]['1sig_0p5t'] = 1 
                hiteff[pdf]['1sig_1p5t'] = 1 
                hiteff[pdf]['1sig_2p5t'] = 1 
                hiteff[pdf]['1sig_3p5t'] = 1
            else:
                hiteff[pdf]['fwhm_1t'] = self.GetCellValue(ws,sheet_fwhm['1tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_2t'] = self.GetCellValue(ws,sheet_fwhm['2tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_3t'] = self.GetCellValue(ws,sheet_fwhm['3tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_fv'] = self.GetCellValue(ws,sheet_fwhm['fvColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_0p5t'] = self.GetCellValue(ws,sheet_fwhm['0p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_1p5t'] = self.GetCellValue(ws,sheet_fwhm['1p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_2p5t'] = self.GetCellValue(ws,sheet_fwhm['2p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['fwhm_3p5t'] = self.GetCellValue(ws,sheet_fwhm['3p5tColumn'],row)*1./hiteff[pdf]['k']

                hiteff[pdf]['3sig_1t'] = self.GetCellValue(ws,sheet_3sig['1tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_2t'] = self.GetCellValue(ws,sheet_3sig['2tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_3t'] = self.GetCellValue(ws,sheet_3sig['3tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_fv'] = self.GetCellValue(ws,sheet_3sig['fvColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_0p5t'] = self.GetCellValue(ws,sheet_3sig['0p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_1p5t'] = self.GetCellValue(ws,sheet_3sig['1p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_2p5t'] = self.GetCellValue(ws,sheet_3sig['2p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['3sig_3p5t'] = self.GetCellValue(ws,sheet_3sig['3p5tColumn'],row)*1./hiteff[pdf]['k']

                hiteff[pdf]['2sig_1t'] = self.GetCellValue(ws,sheet_2sig['1tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_2t'] = self.GetCellValue(ws,sheet_2sig['2tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_3t'] = self.GetCellValue(ws,sheet_2sig['3tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_fv'] = self.GetCellValue(ws,sheet_2sig['fvColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_0p5t'] = self.GetCellValue(ws,sheet_2sig['0p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_1p5t'] = self.GetCellValue(ws,sheet_2sig['1p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_2p5t'] = self.GetCellValue(ws,sheet_2sig['2p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['2sig_3p5t'] = self.GetCellValue(ws,sheet_2sig['3p5tColumn'],row)*1./hiteff[pdf]['k']

                hiteff[pdf]['1sig_1t'] = self.GetCellValue(ws,sheet_1sig['1tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_2t'] = self.GetCellValue(ws,sheet_1sig['2tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_3t'] = self.GetCellValue(ws,sheet_1sig['3tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_fv'] = self.GetCellValue(ws,sheet_1sig['fvColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_0p5t'] = self.GetCellValue(ws,sheet_1sig['0p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_1p5t'] = self.GetCellValue(ws,sheet_1sig['1p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_2p5t'] = self.GetCellValue(ws,sheet_1sig['2p5tColumn'],row)*1./hiteff[pdf]['k']
                hiteff[pdf]['1sig_3p5t'] = self.GetCellValue(ws,sheet_1sig['3p5tColumn'],row)*1./hiteff[pdf]['k']

            self.components[pdf].SetHitEfficiency(hiteff[pdf]['n'],hiteff[pdf]['k'],suffix)
            self.components[pdf].SetHitEfficiencyROI(0,\
                                                     hiteff[pdf]['fwhm_fv'],\
                                                     hiteff[pdf]['fwhm_3t'],\
                                                     hiteff[pdf]['fwhm_2t'],\
                                                     hiteff[pdf]['fwhm_1t'],\
                                                     hiteff[pdf]['fwhm_3p5t'],\
                                                     hiteff[pdf]['fwhm_2p5t'],\
                                                     hiteff[pdf]['fwhm_1p5t'],\
                                                     hiteff[pdf]['fwhm_0p5t'],\
                                                     suffix)
            self.components[pdf].SetHitEfficiencyROI(3,\
                                                     hiteff[pdf]['3sig_fv'],\
                                                     hiteff[pdf]['3sig_3t'],\
                                                     hiteff[pdf]['3sig_2t'],\
                                                     hiteff[pdf]['3sig_1t'],\
                                                     hiteff[pdf]['3sig_3p5t'],\
                                                     hiteff[pdf]['3sig_2p5t'],\
                                                     hiteff[pdf]['3sig_1p5t'],\
                                                     hiteff[pdf]['3sig_0p5t'],\
                                                     suffix)
            self.components[pdf].SetHitEfficiencyROI(2,\
                                                     hiteff[pdf]['2sig_fv'],\
                                                     hiteff[pdf]['2sig_3t'],\
                                                     hiteff[pdf]['2sig_2t'],\
                                                     hiteff[pdf]['2sig_1t'],\
                                                     hiteff[pdf]['2sig_3p5t'],\
                                                     hiteff[pdf]['2sig_2p5t'],\
                                                     hiteff[pdf]['2sig_1p5t'],\
                                                     hiteff[pdf]['2sig_0p5t'],\
                                                     suffix)
            self.components[pdf].SetHitEfficiencyROI(1,\
                                                     hiteff[pdf]['1sig_fv'],\
                                                     hiteff[pdf]['1sig_3t'],\
                                                     hiteff[pdf]['1sig_2t'],\
                                                     hiteff[pdf]['1sig_1t'],\
                                                     hiteff[pdf]['1sig_3p5t'],\
                                                     hiteff[pdf]['1sig_2p5t'],\
                                                     hiteff[pdf]['1sig_1p5t'],\
                                                     hiteff[pdf]['1sig_0p5t'],\
                                                     suffix)
   
        return 
        

    def GetPdfName(self,component,isotope):
        
        if not component == 'LXe':
            for name in self.name_dict:
                if self.name_dict[name] == component:
                    component = name
                    break
        isotope = isotope.replace('-','')
                
        return '%s_%s' % (component,isotope)

    def GetCellName(self,column,row):
        return '%s%d' % (column,row)
    
    def GetCellValue(self,ws,column,row):
        return ws[self.GetCellName(column,row)].value

    def Print(self):

        for pdf in self.components:
            self.components[pdf].Print()

        return

######################################################
############### ROOT TREE WRITER #####################
######################################################
class RootTreeWriter():
    # This class carries info about ROOT tree
    # Adds group information
    # Modify '__init__' appropriately for desired tree

    def __init__(self,outTableName):
        self.filename = outTableName #'../tables/Summary_v68_2016-06-21_0nu.root' #'../tables/Summary_v62_2016-06-04_0nu_tpc_elec.root' #'test_new_tree.root'

        self.pdf_filename_pattern ='/Users/blenardo/Research/nEXO/Sensitivity/sensitivity/histos/nEXO_Histos_%s.root'
    #'/data/data033/exo/software/nEXO_Sensitivity/quick/v5/histos/PNNL/571mm_third/nEXO_Histos_%s.root' # 'individual_histos/nEXO_Histos_%s.root'
        
        self.file = ROOT.TFile.Open(self.filename,'recreate')
        self.tree = ROOT.TTree('ExcelTableValues','Values from Excel Summary Table')
        self.tree.SetName('table') 
        self.table = ROOT.ExcelTableValues()
        
        self.tree.Branch('table',self.table)

        self.groups = {}

        self.groups['Far'] = [ "OuterCryoResin_U238",\
                               "OuterCryoFiber_U238",\
                               "OuterCryoSupportResin_U238",\
                               "OuterCryoSupportFiber_U238",\
                               "InnerCryoResin_U238",\
                               "InnerCryoFiber_U238",\
                               "InnerCryoSupportResin_U238",\
                               "InnerCryoSupportFiber_U238",\
                               "InnerCryoLiner_U238",\
                               "OuterCryoResin_Th232",\
                               "OuterCryoFiber_Th232",\
                               "OuterCryoSupportResin_Th232",\
                               "OuterCryoSupportFiber_Th232",\
                               "InnerCryoResin_Th232",\
                               "InnerCryoFiber_Th232",\
                               "InnerCryoSupportResin_Th232",\
                               "InnerCryoSupportFiber_Th232",\
                               "InnerCryoLiner_Th232",\
                               "OuterCryoResin_Co60",\
                               "OuterCryoFiber_Co60",\
                               "OuterCryoSupportResin_Co60",\
                               "OuterCryoSupportFiber_Co60",\
                               "InnerCryoResin_Co60",\
                               "InnerCryoFiber_Co60",\
                               "InnerCryoSupportResin_Co60",\
                               "InnerCryoSupportFiber_Co60",\
                               "InnerCryoLiner_Co60",\
                               "OuterCryoResin_K40",\
                               "OuterCryoFiber_K40",\
                               "OuterCryoSupportResin_K40",\
                               "OuterCryoSupportFiber_K40",\
                               "InnerCryoResin_K40",\
                               "InnerCryoFiber_K40",\
                               "InnerCryoSupportResin_K40",\
                               "InnerCryoSupportFiber_K40",\
                               "InnerCryoLiner_K40" ]

        group_vessel = ["HFE",\
                        "TPC",\
                        "TPCSupportCone",\
                        "HVTubes",\
                        "HVCables",\
                        "HVFeedthru",\
                        "HVFeedthruCore",\
                        "CalibrationGuideTube1",\
                        "CalibrationGuideTube2" ]

        self.groups['VesselU238'] = ['%s_U238'%(group_comp) for group_comp in group_vessel]
        self.groups['VesselTh232'] = ['%s_Th232'%(group_comp) for group_comp in group_vessel]
        
        group_internal = [ "Cathode",\
                           "Bulge",\
                           "FieldRings",\
                           "SupportSpacers",\
                           "SiPMStaves",\
                           "SiPMElectronics",\
                           "SiPMModule",\
                           "SiPMCables",\
                           "SiPMs",\
                           "ChargeTilesCables",\
                           "ChargeTilesElectronics",\
                           "ChargeTilesSupport",\
                           "ChargeTilesBacking",\
                           "HVPlunger" ]

        self.groups['InternalU238'] = ['%s_U238'%(group_comp) for group_comp in group_internal]
        self.groups['InternalTh232'] = ['%s_Th232'%(group_comp) for group_comp in group_internal]
        
        group_tpc_k40 = ["SupportSpacers","SiPMModule","ChargeTilesBacking"]
        self.groups['FullTpcK40'] = ['%s_K40'%(group_comp) for group_comp in group_tpc_k40]
        
        group_tpc_co60 = ["ChargeTilesCables","ChargeTilesElectronics","ChargeTilesSupport","ChargeTilesBacking","HVPlunger"]
        self.groups['FullTpcCo60'] = ['%s_Co60'%(group_comp) for group_comp in group_tpc_co60]
        
        self.groups['ActiveLXeRn222'] = ["ActiveLXe_Rn222"]
        self.groups['InactiveLXeRn222'] = ["InactiveLXe_Rn222","CathodeRn_Rn222"]
        self.groups['InactiveLXeXe137'] = ["InactiveLXe_Xe137"]
        self.groups['ActiveLXeXe137'] = ["ActiveLXe_Xe137"]
        self.groups['LXeBb2n'] = ["LXe_bb2n"]
        self.groups['LXeBb0n'] = ["LXe_bb0n"]
    
#        self.groups['GhostComponents'] = ["HVCables_K40","HVCables_Co60","HVFeedthruCore_K40"]
        self.groups['GhostComponents'] = ['%s_K40'%(group_comp) for group_comp in group_internal]
        self.groups['GhostComponents'].extend(['%s_Co60'%(group_comp) for group_comp in group_internal])
        self.groups['GhostComponents'].extend(['%s_K40'%(group_comp) for group_comp in group_vessel])
        self.groups['GhostComponents'].extend(['%s_Co60'%(group_comp) for group_comp in group_vessel])
        self.groups['GhostComponents'].extend(['SolderAnode_U238']) 
        self.groups['GhostComponents'].extend(['SolderAnode_Th232']) 
        self.groups['GhostComponents'].extend(['SolderAnode_K40']) 
        self.groups['GhostComponents'].extend(['SolderAnode_Co60']) 
        self.groups['GhostComponents'].extend(['SolderAnode_Ag110m']) 
        self.groups['GhostComponents'].extend(['SolderSiPM_U238']) 
        self.groups['GhostComponents'].extend(['SolderSiPM_Th232']) 
        self.groups['GhostComponents'].extend(['SolderSiPM_K40']) 
        self.groups['GhostComponents'].extend(['SolderSiPM_Co60']) 
        self.groups['GhostComponents'].extend(['SolderSiPM_Ag110m']) 
        self.groups['GhostComponents'].extend(['LXe_B8nu'])
        self.groups['GhostComponents'].extend(['SupportSpacers_Al26'])
        self.groups['GhostComponents'].extend(['FieldRings_Cs137'])

    def FindGroup(self, pdf):

        for name in self.groups:
            group = self.groups[name]
            for component in group:
                if pdf == component:
                    return name

        print '***** CANNOT FIND GROUP FOR %s *****' % (pdf)
        return None

    def Fill(self, values):
        print 'Filling tree...'

        values.SetGroup(self.FindGroup(values.fPdf))
        values.SetFileName(self.pdf_filename_pattern % values.fPdf)
        #if values.fIsotope == 'bb0n':
        #    values.fFileName = values.fFileName.replace('resol0.005','resol0.010')
        
        self.table.Copy(values)
        self.table.Print()

        self.tree.Fill()

    def Write(self):
        print 'Writing tree into file...'
        self.file.cd()
        self.tree.Write()

    def CloseFile(self):
        print 'Closing file...', self.file.GetName()
        self.file.Close()        

######################################################
############# TABLE TO TREE CONVERTER ################
######################################################
class Excel2RootConverter():
    # This class converts the table into tree

    def __init__(self, excelTable, rootTree):
        self.table = excelTable
        self.tree = rootTree

    def Run(self):
        print 'Converting table to tree...'
        for pdf in self.table.components:
            self.tree.Fill(self.table.components[pdf])
        self.tree.Write()
        self.tree.CloseFile()

######################################################
####################  MAIN  ##########################
######################################################
import sys

if __name__ == "__main__":
   
    
    if len(sys.argv) == 3:
        inTableName = sys.argv[1] # '../tables/Summary_v68_2016-06-21_0nu.xlsx'
        outTableName = sys.argv[2] # '../tables/Summary_v68_2016-06-21_0nu.root'
    else:
        sys.exit('\nERROR: Must enter input and output table names...\n')

    start_time = time.time()
    print('Script started at {}'.format(start_time))

    excelTable = ExcelTableReader(inTableName) #excelTable.Print()
    rootTree = RootTreeWriter(outTableName)

 
    table2tree = Excel2RootConverter(excelTable,rootTree)
    table2tree.Run()

    end_time = time.time()
    print('Elapsed time = {} seconds ({} minutes).'.format( end_time-start_time, (end_time-start_time)/60. ) )
