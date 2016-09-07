
from openpyxl import load_workbook

#### General options ####
number_components = 72
quantile = 1.64

excel_filename = '../tables/Summary_01042016_Si_Cu_v2.xlsx'
wb = load_workbook(filename=excel_filename,read_only=True,data_only=True)

specific_activity_sheet = {'name': 'SpecificActivities', 'startRow': 2, 'componentColumn': 'A', 'isotopeColumn': 'B'}
counts_sheet = {'name': '%s_ExpectedCounts', 'startRow': 6, 'componentColumn': 'A', 'isotopeColumn': 'B', 'cvColumn': {'SS': 'AR', 'MS': 'AH'}, 'errorColumn': {'SS': 'AS', 'MS': 'AI'}}

name_dict = {}
name_dict["OuterCryo"] = 'Outer Cryostat'
name_dict["InnerCryo"] = 'Inner Cryostat'
name_dict["InnerCryoLiner"] = 'Inner Cryostat Liner'
name_dict["HFE"] = 'HFE'
name_dict["TPC"] = 'TPC Vessel'
name_dict["Cathode"] = 'Cathode'
name_dict["Bulge"] = 'Bulge'
name_dict["FieldRing"] = 'Field Ring'
name_dict["SupportLeg"] = 'Support Leg'
name_dict["SupportSpacer"] = 'Support Spacer'
name_dict["SiPMSupport"] = 'SiPM Support'
name_dict["SiPMQuartz"] = 'SiPM Quartz'
name_dict["SiPMElectronics"] = 'SiPM Electronics'
name_dict["SiPMGlue"] = 'SiPM Glue'
name_dict["SiPMCables"] = 'SiPM Cables'
name_dict["SiPM"] = 'SiPM'
name_dict["ChargeModuleCables"] = 'Charge Module Cables'
name_dict["ChargeModuleChip"] = 'Charge Module Chip'
name_dict["ChargeModuleGlue"] = 'Charge Module Glue'
name_dict["ChargeModuleSupport"] = 'Charge Module Support'
name_dict["ChargeModuleBacking"] = 'Charge Module Backing'
name_dict["LXe"] = 'LXe'
name_dict["bb2n"] = 'LXe'
name_dict["ActiveLXe"] = 'Active LXe'
name_dict["InactiveLXe"] = 'Inactive LXe'

#### Grouping options ####

groups = {}

groups['Far'] = ["OuterCryo_Co60", "OuterCryo_Th232", "OuterCryo_U238","OuterCryo_K40","InnerCryo_Co60", "InnerCryo_Th232", "InnerCryo_U238","InnerCryo_K40","InnerCryoLiner_Co60", "InnerCryoLiner_Th232", "InnerCryoLiner_U238"]

group_vessel = ["HFE","TPC"]
groups['VesselU238'] = ['%s_U238'%(group_comp) for group_comp in group_vessel]
groups['VesselTh232'] = ['%s_Th232'%(group_comp) for group_comp in group_vessel]

group_internal = ["Cathode","Bulge","FieldRing","SupportLeg","SupportSpacer","SiPMSupport","SiPMQuartz","SiPMElectronics","SiPMGlue","SiPMCables","SiPM","ChargeModuleCables","ChargeModuleChip","ChargeModuleGlue","ChargeModuleSupport","ChargeModuleBacking"]
groups['InternalU238'] = ['%s_U238'%(group_comp) for group_comp in group_internal]
groups['InternalTh232'] = ['%s_Th232'%(group_comp) for group_comp in group_internal]

group_tpc_k40 = ["TPC","Cathode","Bulge","FieldRing","SupportLeg","SupportSpacer","SiPMSupport","SiPMElectronics","SiPMGlue","SiPM","ChargeModuleChip","ChargeModuleGlue","ChargeModuleSupport"]
groups['FullTpcK40'] = ['%s_K40'%(group_comp) for group_comp in group_tpc_k40]

group_tpc_co60 = ["TPC","Cathode","Bulge","FieldRing","SiPMSupport","SiPMElectronics","ChargeModuleChip","ChargeModuleSupport"]
groups['FullTpcCo60'] = ['%s_Co60'%(group_comp) for group_comp in group_tpc_co60]

groups['LXeRn222'] = ["ActiveLXe_Rn222","InactiveLXe_Rn222"]
groups['LXeXe137'] = ["LXe_Xe137"]
groups['LXeBb2n'] = ["LXe_bb2n"]
#groups['lxeBb0n'] = ["LXe_bb0n"]

#### Auxiliary functions ####

def GetPdfName(component,isotope):

    if not component == 'LXe':
        for name in name_dict:
            if name_dict[name] == component:
                component = name
                break
    isotope = isotope.replace('-','')

    return '%s_%s' % (component,isotope)

#### Table reading functions ####

def GetCellName(column,row):
    return '%s%d' % (column,row)

def GetCellValue(ws,column,row):
    return ws[GetCellName(column,row)].value

def ReadCounts(suffix):

    global wb

    ws = wb.get_sheet_by_name(counts_sheet['name']%(suffix))

    counts = {}
    start_row = counts_sheet['startRow']
    end_row = counts_sheet['startRow']+number_components
    for row in range(start_row,end_row):
        component = GetCellValue(ws,counts_sheet['componentColumn'],row)
        isotope = GetCellValue(ws,counts_sheet['isotopeColumn'],row)
        pdf = GetPdfName(component,isotope)
        print row, '/', end_row, ':', component, isotope, pdf
        counts[pdf] = {'cv':GetCellValue(ws,counts_sheet['cvColumn'][suffix],row),'error':GetCellValue(ws,counts_sheet['errorColumn'][suffix],row)}
        counts[pdf]['ul'] = (quantile*counts[pdf]['error'],counts[pdf]['cv'] + quantile*counts[pdf]['error'])[counts[pdf]['cv'] > 0]
    
    return counts

#### Printing functions ####

def PrintComponents():
    global wb

    ws = wb.get_sheet_by_name(specific_activity_sheet['name'])
    for row in range(specific_activity_sheet['startRow'],specific_activity_sheet['startRow']+number_components):
        component = GetCellValue(ws,specific_activity_sheet['componentColumn'],row)
        isotope = GetCellValue(ws,specific_activity_sheet['isotopeColumn'],row)
        pdf = GetPdfName(component,isotope)
        #print 'fAllComponents.push_back("%s");'%(pdf),
        print 'fComponents.insert(std::make_pair("%s",ExcelTable("%s","%s")));'%(pdf,component,isotope),
    print '\n'
        
def PrintGroups():

    for name in groups:
        print 'fGroups.insert(std::make_pair("%s",std::vector<TString>()));' % (name)
        group = groups[name]
        for component in group:
            print 'fGroups["%s"].push_back("%s");' % (name,component),
        print '\n'

def PrintCounts(suffix):
    
    counts = ReadCounts(suffix)
    for pdf in counts:
        table = counts[pdf]
        print 'fComponents["%s"].SetExpectedCounts(%.2e,%.2e,%.2e,"%s");' % (pdf,table['cv'],table['error'],table['ul'],suffix),
    print '\n'
    
if __name__ == "__main__":

    #PrintComponents()
    #PrintGroups()
    #PrintCounts('SS')
    PrintCounts('MS')
