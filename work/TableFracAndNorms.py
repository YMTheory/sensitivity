

from openpyxl import load_workbook

#### General options ####
excel_filename = '../tables/Summary_01042016_Si_Cu_v2.xlsx'

ss_sheet_name = 'SS_ExpectedCounts'
ms_sheet_name = 'MS_ExpectedCounts'

component_column = 'A'
isotope_column = 'B'
start_row = 6
end_row = 77

cv_column_ss = 'AR' #'AP' # 'AQ'
error_column_ss = 'AS' #'AQ' #'AR'

cv_column_ms = 'AH' #'AP' # 'AQ'
error_column_ms = 'AI' #'AQ' #'AR'

quantile = 1.64

name_dict = {}
name_dict["OuterCryo"] = 'Outer Cryostat'
name_dict["InnerCryo"] = 'Inner Cryostat'
name_dict["InnerCryoLiner"] = 'Inner Cryostat Liner'
name_dict["HFE"] = 'HFE'
name_dict["TPC"] = 'TPC Vessel'
name_dict["Cathode"] = 'Cathode'
name_dict["Bulge"] = 'Bulge'
name_dict["FieldRing"] = 'Field Ring'
name_dict["SupportLeg"] = 'Support Leg'
name_dict["SupportSpacer"] = 'Support Spacer'
name_dict["SiPMSupport"] = 'SiPM Support'
name_dict["SiPMQuartz"] = 'SiPM Quartz'
name_dict["SiPMElectronics"] = 'SiPM Electronics'
name_dict["SiPMGlue"] = 'SiPM Glue'
name_dict["SiPMCables"] = 'SiPM Cables'
name_dict["SiPM"] = 'SiPM'
name_dict["ChargeModuleCables"] = 'Charge Module Cables'
name_dict["ChargeModuleChip"] = 'Charge Module Chip'
name_dict["ChargeModuleGlue"] = 'Charge Module Glue'
name_dict["ChargeModuleSupport"] = 'Charge Module Support'
name_dict["ChargeModuleBacking"] = 'Charge Module Backing'
name_dict["LXe"] = 'LXe'
name_dict["bb2n"] = 'LXe'
name_dict["ActiveLXe"] = 'Active LXe'
name_dict["InactiveLXe"] = 'Inactive LXe'


#### Grouping options ####
groups = {}

groups['far'] = ["OuterCryo_Co60", "OuterCryo_Th232", "OuterCryo_U238","OuterCryo_K40","InnerCryo_Co60", "InnerCryo_Th232", "InnerCryo_U238","InnerCryo_K40","InnerCryoLiner_Co60", "InnerCryoLiner_Th232", "InnerCryoLiner_U238"]

group_vessel = ["HFE","TPC"]
groups['vesselU238'] = ['%s_U238'%(group_comp) for group_comp in group_vessel]
groups['vesselTh232'] = ['%s_Th232'%(group_comp) for group_comp in group_vessel]

group_internal = ["Cathode","Bulge","FieldRing","SupportLeg","SupportSpacer","SiPMSupport","SiPMQuartz","SiPMElectronics","SiPMGlue","SiPMCables","SiPM","ChargeModuleCables","ChargeModuleChip","ChargeModuleGlue","ChargeModuleSupport","ChargeModuleBacking"]
groups['internalU238'] = ['%s_U238'%(group_comp) for group_comp in group_internal]
groups['internalTh232'] = ['%s_Th232'%(group_comp) for group_comp in group_internal]

group_tpc_k40 = ["TPC","Cathode","Bulge","FieldRing","SupportLeg","SupportSpacer","SiPMSupport","SiPMElectronics","SiPMGlue","SiPM","ChargeModuleChip","ChargeModuleGlue","ChargeModuleSupport"]
groups['tpcK40'] = ['%s_K40'%(group_comp) for group_comp in group_tpc_k40]

group_tpc_co60 = ["TPC","Cathode","Bulge","FieldRing","SiPMSupport","SiPMElectronics","ChargeModuleChip","ChargeModuleSupport"]
groups['tpcCo60'] = ['%s_Co60'%(group_comp) for group_comp in group_tpc_co60]

groups['lxeRn222'] = ["ActiveLXe_Rn222","InactiveLXe_Rn222"]
groups['lxeXe137'] = ["LXe_Xe137"]

#### Code ####
wb = load_workbook(filename=excel_filename,read_only=True,data_only=True)

ws_ss = wb.get_sheet_by_name(ss_sheet_name)
ws_ms = wb.get_sheet_by_name(ms_sheet_name)

def read_counts(ws,cv_column,error_column):

    print 'Reading counts...'

    counts = {}
    for row in range(start_row,end_row+1):
        print row, '/', end_row
        cell_name = '%s%d' % (component_column,row)
        table_name = ws[cell_name].value
        if not table_name in counts:
            counts[table_name] = {}
        cell_isotope = '%s%d' % (isotope_column,row)
        table_isotope = ws[cell_isotope].value.replace('-','')
        counts[table_name][table_isotope] = {'cv':0,'error':0,'ul':0}

        cell_cv = '%s%d' % (cv_column,row)
        cell_error = '%s%d' % (error_column,row)
        table_cv, table_error = ws[cell_cv].value, ws[cell_error].value

        counts[table_name][table_isotope]['cv'] = table_cv
        counts[table_name][table_isotope]['error'] = table_error
        counts[table_name][table_isotope]['ul'] =  (quantile*table_error,table_cv + quantile*table_error)[table_cv > 0]

    return counts

counts_ss = read_counts(ws_ss,cv_column_ss,error_column_ss)
counts_ms = read_counts(ws_ms,cv_column_ms,error_column_ms)

def print_group_names(group,name):

    cppCodeLine = 'TString %sComponents[%d] = {' % (name,len(group))
    for component in group:
        cppCodeLine += '"%s",' % (component)
    cppCodeLine = cppCodeLine[:-1]
    cppCodeLine += '}'

    print cppCodeLine   

def print_group_fractions(counts,group,name,suffix):

    fractions = []
    sum_cv = 0.
    sumsq_error = 0.

    for component in group:
        group_name = component[:component.rfind('_')]
        group_isotope = component[component.rfind('_')+1:] 
        group_ul = counts[name_dict[group_name]][group_isotope]['ul']

        sum_cv += counts[name_dict[group_name]][group_isotope]['cv']
        sumsq_error += counts[name_dict[group_name]][group_isotope]['error']**2

        fractions.append(group_ul)
    
    print_fractions = [float('%.2f' % frac) for frac in fractions]
    sum_error = sumsq_error**0.5
    sum_ul = (quantile*sum_error,sum_cv + quantile*sum_error)[sum_cv > 0]
    #print len(print_fractions), print_fractions, sum_ul  

    cppCodeLine = 'Double_t %sFractions%s[%d] = {' % (name,suffix,len(group))
    for fraction in fractions:
        cppCodeLine += '%.2f,' % (fraction)
    cppCodeLine = cppCodeLine[:-1]
    cppCodeLine += '}'

    print cppCodeLine   
    

for name in groups:
    group = groups[name]
    print name

    print_group_names(group,name)

    print_group_fractions(counts_ss,group,name,'SS')
    print_group_fractions(counts_ms,group,name,'MS')

#for row in range(start_row,end_row+1):
#    cell_name = '%s%d' % (component_column,row)
#    cell_isotope = '%s%d' % (isotope_column,row)
#    cell_cv = '%s%d' % (cv_column,row)
#    cell_error = '%s%d' % (error_column,row)
#    print row, ws_ss[cell_name].value, ws_ss[cell_isotope].value, ws_ss[cell_cv].value, ws_ss[cell_error].value


#  TString farComponents[11] = {"OuterCryo_Co60", "OuterCryo_Th232", "OuterCryo_U238","OuterCryo_K40","InnerCryo_Co60", "InnerCryo_Th232", "InnerCryo_U238","InnerCryo_K40","InnerCryoLiner_Co60", "InnerCryoLiner_Th232", "InnerCryoLiner_U238"};
#  Double_t farFractionsSS[11] = {3.28, 26.58, 10.45, 51.68, 1.31, 9.96, 4.52, 21.53, 0.25, 3.35, 0.49};
#  Double_t farFractionsMS[11] = {5.01, 78.76, 27.76, 108.23, 2.02, 30.94, 10.92, 44.56, 0.39, 10.4, 1.18};

